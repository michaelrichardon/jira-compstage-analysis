"""
Jira CompStage & Story Analyse
================================
Spalten (Excel):
  Team Mapping | Key | Zusammenfassung | Functional Release | Release | Not-Testable |
  Defects offen | Defects gesamt | Status

CompStage-Zeilen: alle Spalten befüllt (fett, blau)
Story-Zeilen:     Key + Zusammenfassung + Release + Not-Testable + Defects + Status

Delta-Modus:
  Beim zweiten+ Run wird die bestehende DB mit dem neuen Abruf verglichen.
  Nur geänderte/neue Issues werden in der DB aktualisiert.
  Eine Tabelle "delta_log" protokolliert jede Änderung mit Timestamp.

Konfiguration via .env:
  JIRA_BASE_URL, JIRA_PAT, JIRA_PROJECT, JIRA_FEATURE_TEAM_FIELD,
  JIRA_SSL_VERIFY, JIRA_DB_PATH, OUTPUT_EXCEL
"""

import os
import sys
import sqlite3
import urllib3
import requests
from datetime import datetime, timezone
import html as html_mod
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# .env laden
# ---------------------------------------------------------------------------
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ---------------------------------------------------------------------------
# Konfiguration
# ---------------------------------------------------------------------------
JIRA_BASE_URL      = os.environ.get("JIRA_BASE_URL",           "https://jira.local.wmgruppe.de").rstrip("/")
JIRA_PAT           = os.environ.get("JIRA_PAT",                "")
JIRA_PROJECT       = os.environ.get("JIRA_PROJECT",            "FONTUS")
FEATURE_TEAM_FIELD = os.environ.get("JIRA_FEATURE_TEAM_FIELD", "customfield_11101")
SSL_VERIFY         = os.environ.get("JIRA_SSL_VERIFY",         "false").strip().lower() not in ("false", "0", "no")
DB_PATH            = os.environ.get("JIRA_DB_PATH",            "fontus_CS_analysis.db")
_ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
_output_dir  = os.environ.get("OUTPUT_DIR", r"C:\Projekte\JiraCompStage analysis\Ergebnisdateien")
OUTPUT_EXCEL = os.environ.get("OUTPUT_EXCEL",
               os.path.join(_output_dir, f"compstage_story_analysis_{_ts}.xlsx"))

FUNC_RELEASE_FIELD = "customfield_11313"   # Functional Release – nur CompStage

TRACKING_SHEET_URL = os.environ.get(
    "TRACKING_SHEET_URL",
    "https://confluence.local.wmgruppe.de/download/attachments/313855600/R3x_OverallTrackingSheet.xlsx",
)

CONFLUENCE_BASE_URL    = os.environ.get("CONFLUENCE_BASE_URL",    "https://confluence.local.wmgruppe.de").rstrip("/")
CONFLUENCE_PAT         = os.environ.get("CONFLUENCE_PAT",         "")
CONFLUENCE_PARENT_ID   = os.environ.get("CONFLUENCE_PARENT_PAGE_ID", "328092016")
CONFLUENCE_SPACE       = os.environ.get("CONFLUENCE_SPACE",       "FONTUSIBM")

# Confluence HTTP-Session (lazy – wird erst bei Bedarf verwendet)
_cf_session: requests.Session = None  # wird in export_confluence initialisiert

if not JIRA_PAT:
    print("FEHLER: JIRA_PAT ist nicht gesetzt.")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Tracking-Sheet: Release-Mapping aus Confluence
# ---------------------------------------------------------------------------
def load_tracking_releases() -> dict[str, str]:
    """
    Lädt R3x_OverallTrackingSheet.xlsx von Confluence.
    Liest Spalte A (Release) und Spalte B (Issue-ID = CompStage-Key).
    Gibt dict {compstage_key → release} zurück.
    Zeilen mit n/a oder leerem Issue-Key werden übersprungen.
    """
    if not TRACKING_SHEET_URL:
        return {}
    if not CONFLUENCE_PAT:
        print("  Tracking-Sheet: CONFLUENCE_PAT nicht gesetzt – übersprungen.")
        return {}
    try:
        import io as _io
        import openpyxl as _xl
        print(f"  Lade Tracking-Sheet von Confluence …")
        s = requests.Session()
        s.headers.update({"Authorization": f"Bearer {CONFLUENCE_PAT}"})
        s.verify = SSL_VERIFY
        resp = s.get(TRACKING_SHEET_URL, stream=True)
        resp.raise_for_status()
        wb   = _xl.load_workbook(_io.BytesIO(resp.content), read_only=True, data_only=True)
        ws   = wb.active
        mapping: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            release  = str(row[0]).strip() if row[0] else ""
            issue_id = str(row[1]).strip() if row[1] else ""
            if not issue_id or issue_id.lower() in ("n/a", "none", ""):
                continue
            if not release or release.lower() in ("n/a", "none", ""):
                continue
            mapping[issue_id] = release
        wb.close()
        print(f"  Tracking-Sheet: {len(mapping)} Einträge geladen.")
        return mapping
    except Exception as e:
        print(f"  Tracking-Sheet FEHLER: {e}")
        return {}


# ---------------------------------------------------------------------------
# Feature Teams & Mapping
# ---------------------------------------------------------------------------
FEATURE_TEAMS = [
    "DL-01", "DL-02", "DL-GenAI", "DLH", "DLH Source",
    "DL-Output", "UI&S", "User Interfaces", "T2-Team",
    "Rule Engine", "Rule Validation",
]

TEAM_MAPPING = {
    "DL-01":           "DLH",
    "DL-02":           "DLH",
    "DL-GenAI":        "DLH",
    "DLH":             "DLH",
    "DLH Source":      "DLH",
    "DL-Output":       "DLH",
    "UI&S":            "User Interfaces",
    "User Interfaces": "User Interfaces",
    "T2-Team":         "Rule Engine",
    "Rule Engine":     "Rule Engine",
    "Rule Validation": "Rule Validation",
}

# ---------------------------------------------------------------------------
# Release-Labels & Defect-Konfiguration
# ---------------------------------------------------------------------------
RELEASE_LABELS = [
    "R2.1", "R2.2", "R2.3",
    "R3.1", "R3.2", "R3.3", "R3.4", "R3.5", "R3.6", "R3.7", "R3.8",
    "R3.9", "R3.10", "R3.11", "R3.12", "R3.13", "R3.14", "R3.15", "R3.16", "R3.17",
]
RELEASE_SET = set(RELEASE_LABELS)

NOT_TESTABLE_LABEL = "Not-Testable"

# Status die als "offen" gelten (alles AUSSER diesen → offen)
DEFECT_CLOSED_STATUSES = {"Done", "Closed", "Rejected"}

# ---------------------------------------------------------------------------
# Status-Konsistenzprüfung CompStage ↔ Stories
#
# Kat-Index-CS:
#   1 = DEV        → Stories müssen Kat-Index >= 1 haben
#   2 = K-Test     → Stories müssen Kat-Index >= 2 haben
#   3 = INT / UAT  → Stories müssen Kat-Index >= 3 (= CLOSED) haben
#  99 = CLOSED     → Stories müssen Kat-Index >= 3 (= CLOSED) haben
# 999 = Ignore     → keine Prüfung
#
# Kat-Index-Story:
#   0   = Ignore (Pending, Rejected)
#   1   = DEV
#   2   = TEST
#   3   = CLOSED (Done, Closed)
# 999   = Ignore (Deprecated)
#
# Not-Testable Stories werden ignoriert.
# Stories mit Kat-Index 0 oder 999 werden ignoriert.
# ---------------------------------------------------------------------------

# (Status → (Einzel-Index, Kategorie, Kat-Index))
CS_STATUS_MAP = {
    "To Do":                  ( 1, "DEV",    1),
    "In Review":              ( 2, "DEV",    1),
    "In Specification":       ( 3, "DEV",    1),
    "Ready for Design":       ( 4, "DEV",    1),
    "In Design":              ( 5, "DEV",    1),
    "In Test Creation":       ( 6, "DEV",    1),
    "Ready for Development":  ( 7, "DEV",    1),
    "In Progress":            ( 8, "DEV",    1),
    "Ready for CompTest":     ( 9, "K-Test", 2),
    "In CompTest":            (10, "K-Test", 2),
    "Ready for INT":          (11, "INT",    3),
    "In INT":                 (12, "INT",    3),
    "INT Tested":             (13, "INT",    3),
    "In UAT":                 (14, "UAT",    3),
    "UAT Tested":             (15, "UAT",    3),
    "Closed":                 (16, "CLOSED", 99),
    "Pending":                ( 0, "Ignore", 999),
    "Rejected":               (-1, "Ignore", 999),
}

STORY_STATUS_MAP = {
    "To Do":                  ( 1, "DEV",    1),
    "In Specification":       ( 2, "DEV",    1),
    "Ready for Estimation":   ( 3, "DEV",    1),
    "Reopened":               ( 4, "DEV",    1),
    "Ready for Development":  ( 5, "DEV",    1),
    "In Progress":            ( 6, "DEV",    1),
    "In Review":              ( 7, "DEV",    1),
    "Ready For Deployment":   ( 8, "DEV",    1),
    "Ready for Test":         ( 9, "TEST",   2),
    "In Test":                (10, "TEST",   2),
    "Done":                   (11, "CLOSED", 3),
    "Closed":                 (12, "CLOSED", 3),
    "Deprecated":             (-1, "Ignore", 999),
    "Pending":                (-1, "Ignore", 0),
    "Rejected":               ( 0, "Ignore", 0),
}

# Minimaler Story-Kat-Index je CS-Kat-Index
# CS 99 (CLOSED) verhält sich wie CS 3 (INT/UAT): alle Stories müssen closed sein
CS_MIN_STORY_KAT = {
    1:   1,    # DEV   → Story >= DEV
    2:   2,    # K-Test → Story >= TEST
    3:   3,    # INT/UAT → Story muss CLOSED sein
    99:  3,    # CLOSED  → Story muss ebenfalls CLOSED sein
    999: None, # Ignore  → keine Prüfung
}


def cs_kat_index(status: str) -> int:
    return CS_STATUS_MAP.get(status, (0, "?", 999))[2]

def story_kat_index(status: str) -> int:
    return STORY_STATUS_MAP.get(status, (0, "?", 999))[2]


def check_cs_consistency(cs_status: str, story_rows: list) -> tuple[str, set]:
    """
    Prüft ob alle relevanten Stories den Mindest-Kat-Index für den CS-Status erfüllen.
    Ignoriert: Not-Testable Stories, Stories mit Kat-Index 0 (Pending/Rejected) oder 999 (Deprecated).
    Gibt (cs_flag, bad_story_keys) zurück:
      cs_flag       = 'X' wenn mindestens eine Story den Mindest-Index unterschreitet
      bad_story_keys = Set der Story-Keys die die Inkonsistenz verursachen
    """
    cs_ki     = cs_kat_index(cs_status)
    min_st_ki = CS_MIN_STORY_KAT.get(cs_ki)
    bad_keys: set[str] = set()

    if min_st_ki is None:          # CS = Ignore → nicht prüfen
        return "", bad_keys

    for st_row in story_rows:
        if st_row.get("not_testable") == "X":
            continue
        st_ki = story_kat_index(st_row["status"])
        if st_ki in (0, 999):      # Pending / Rejected / Deprecated → ignorieren
            continue
        if st_ki < min_st_ki:
            bad_keys.add(st_row["key"])

    return ("X" if bad_keys else ""), bad_keys


def extract_labels(issue: dict) -> list:
    return issue.get("fields", {}).get("labels") or []

def extract_release(issue: dict) -> str:
    hits = [l for l in extract_labels(issue) if l in RELEASE_SET]
    return ", ".join(hits)

def extract_not_testable(issue: dict) -> str:
    return "X" if NOT_TESTABLE_LABEL in extract_labels(issue) else ""

# ---------------------------------------------------------------------------
# HTTP-Session mit PAT-Auth
# ---------------------------------------------------------------------------
session = requests.Session()
session.headers.update({
    "Authorization": f"Bearer {JIRA_PAT}",
    "Accept":        "application/json",
    "Content-Type":  "application/json",
})
session.verify = SSL_VERIFY

# ---------------------------------------------------------------------------
# Jira REST-Hilfsfunktionen
# ---------------------------------------------------------------------------
def jql_search(jql: str, fields: list, max_results: int = 2000) -> list:
    url    = f"{JIRA_BASE_URL}/rest/api/2/search"
    issues = []
    start  = 0
    while True:
        params = {
            "jql":        jql,
            "fields":     ",".join(fields),
            "maxResults": min(100, max_results - len(issues)),
            "startAt":    start,
        }
        resp = session.get(url, params=params)
        resp.raise_for_status()
        data  = resp.json()
        batch = data.get("issues", [])
        issues.extend(batch)
        start += len(batch)
        if start >= data.get("total", 0) or len(issues) >= max_results or not batch:
            break
    return issues


def get_issue(key: str, fields: list) -> dict:
    url  = f"{JIRA_BASE_URL}/rest/api/2/issue/{key}"
    resp = session.get(url, params={"fields": ",".join(fields)})
    resp.raise_for_status()
    return resp.json()


def f_summary(issue: dict) -> str:
    return issue.get("fields", {}).get("summary", "")

def f_status(issue: dict) -> str:
    return issue.get("fields", {}).get("status", {}).get("name", "")

def f_issuetype(issue: dict) -> str:
    return issue.get("fields", {}).get("issuetype", {}).get("name", "")

def f_feature_team(issue: dict) -> str | None:
    raw = issue.get("fields", {}).get(FEATURE_TEAM_FIELD)
    if raw is None:
        return None
    if isinstance(raw, dict):
        return raw.get("value") or raw.get("name")
    return str(raw)

def f_functional_release(issue: dict) -> str:
    raw = issue.get("fields", {}).get(FUNC_RELEASE_FIELD)
    if raw is None:
        return ""
    if isinstance(raw, list):
        parts = []
        for item in raw:
            if isinstance(item, dict):
                parts.append(item.get("value") or item.get("name") or str(item))
            else:
                parts.append(str(item))
        return ", ".join(parts)
    if isinstance(raw, dict):
        return raw.get("value") or raw.get("name") or ""
    return str(raw)

# ---------------------------------------------------------------------------
# Defect-Zählung für eine Story
# ---------------------------------------------------------------------------
def get_defect_counts(story_key: str) -> tuple[int, int]:
    """
    Gibt (defects_offen, defects_gesamt) zurück.
    Liest alle issuelinks der Story und filtert auf issuetype = Bug (= Defect in diesem Projekt).
    linkedIssues()-JQL wird nicht verwendet (400-Fehler auf Jira Server).
    """
    found: dict[str, str] = {}   # key → status
    try:
        full = get_issue(story_key, fields=["issuelinks"])
        for link in full.get("fields", {}).get("issuelinks", []):
            for direction in ("outwardIssue", "inwardIssue"):
                linked = link.get(direction)
                if linked and f_issuetype(linked) in ("Defect", "Bug"):
                    status = linked.get("fields", {}).get("status", {}).get("name", "")
                    found[linked["key"]] = status
    except Exception as e:
        print(f"      [Defects skip {story_key}]: {e}")

    gesamt = len(found)
    offen  = sum(1 for s in found.values() if s not in DEFECT_CLOSED_STATUSES)
    return offen, gesamt

# ---------------------------------------------------------------------------
# Stories zu einer CompStage – 4-Pfad-Strategie
# ---------------------------------------------------------------------------
STORY_FIELDS = ["summary", "status", "labels", FEATURE_TEAM_FIELD]

def get_stories_for_compstage(cs: dict) -> list:
    cs_key = cs["key"]
    found: dict[str, dict] = {}

    # Pfad 1: issuelinks
    for link in cs.get("fields", {}).get("issuelinks", []):
        for direction in ("outwardIssue", "inwardIssue"):
            linked = link.get(direction)
            if linked and f_issuetype(linked) == "Story":
                found[linked["key"]] = linked

    # Pfad 2: subtasks
    for sub in cs.get("fields", {}).get("subtasks", []):
        if f_issuetype(sub) == "Story":
            found[sub["key"]] = sub

    # Pfad 3: JQL parent
    try:
        for ci in jql_search(
            f'parent = "{cs_key}" AND issuetype = Story ORDER BY key ASC',
            fields=STORY_FIELDS,
        ):
            found[ci["key"]] = ci
    except Exception as e:
        print(f"    [Pfad 3 skip {cs_key}]: {e}")

    if not found:
        return []

    stories = []
    for key in sorted(found.keys()):
        try:
            stories.append(get_issue(key, fields=STORY_FIELDS))
        except Exception as e:
            print(f"    Warnung – Story {key}: {e}")
    return stories

# ---------------------------------------------------------------------------
# Hauptanalyse
# ---------------------------------------------------------------------------
CS_FIELDS = [
    "summary", "status", "labels",
    FEATURE_TEAM_FIELD, FUNC_RELEASE_FIELD,
    "issuelinks", "subtasks", "issuetype",
]

def build_rows(tracking_map: dict | None = None) -> list:
    if tracking_map is None:
        tracking_map = {}
    teams_jql = ", ".join(f'"{t}"' for t in FEATURE_TEAMS)
    jql = (
        f'project = "{JIRA_PROJECT}" '
        f'AND issuetype = "CompStage" '
        f'AND cf[11101] in ({teams_jql}) '
        f'ORDER BY key ASC'
    )
    print(f"JQL:\n  {jql}\n")

    compstages = jql_search(jql, fields=CS_FIELDS)
    print(f"CompStages gefunden: {len(compstages)}\n")

    compstages.sort(key=lambda c: (
        TEAM_MAPPING.get(f_feature_team(c), f_feature_team(c) or ""),
        c["key"]
    ))

    rows = []
    for cs in compstages:
        raw_team  = f_feature_team(cs)
        team_name = TEAM_MAPPING.get(raw_team, raw_team or "Unbekannt")

        cs_status = f_status(cs)
        stories   = get_stories_for_compstage(cs)
        print(f"  {cs['key']:20s} [{team_name:20s}]  → {len(stories)} Stories")

        # Story-Rows erst aufbauen …
        story_rows = []
        for st in stories:
            d_open, d_total = get_defect_counts(st["key"])
            print(f"    {st['key']:18s}  Bugs: {d_open} offen / {d_total} gesamt")
            st_status = f_status(st)
            not_test  = extract_not_testable(st)
            story_rows.append({
                "row_type":           "story",
                "team_mapped":        "",
                "key":                st["key"],
                "summary":            f_summary(st),
                "functional_release": "",
                "release":            extract_release(st),
                "tracking_release":   "",   # nur CompStage
                "not_testable":       not_test,
                "defects_open":       d_open,
                "defects_total":      d_total,
                "status":             st_status,
                "inconsistent":       "X" if st_status in ("Done", "Closed") and d_open > 0 else "",
                "cs_inconsistent":    "",   # wird auf CS-Zeile gesetzt
                "parent_key":         cs["key"],
            })

        # … dann CS-Konsistenz prüfen (braucht vollständige story_rows)
        cs_incons, bad_keys = check_cs_consistency(cs_status, story_rows)
        if cs_incons:
            print(f"  → CS {cs['key']} INKONSISTENT (Status={cs_status}, "
                  f"{len(bad_keys)} Story/ies zu früh)")

        # Stories die CS-Inkonsistenz verursachen mit "x" markieren
        for st_row in story_rows:
            if st_row["key"] in bad_keys:
                st_row["cs_inconsistent"] = "x"

        rows.append({
            "row_type":           "compstage",
            "team_mapped":        team_name,
            "key":                cs["key"],
            "summary":            f_summary(cs),
            "functional_release": f_functional_release(cs),
            "release":            extract_release(cs),
            "tracking_release":   tracking_map.get(cs["key"], ""),
            "not_testable":       extract_not_testable(cs),
            "defects_open":       "",
            "defects_total":      "",
            "status":             cs_status,
            "inconsistent":       cs_incons,
            "cs_inconsistent":    cs_incons,
            "parent_key":         None,
        })
        rows.extend(story_rows)

    return rows

# ---------------------------------------------------------------------------
# Excel-Export
# ---------------------------------------------------------------------------
STATUS_COLORS = {
    "To Do":       "DDDDDD",
    "In Progress": "DEEBFF",
    "In Review":   "FFF0B3",
    "Done":        "E3FCEF",
    "Closed":      "E3FCEF",
    "Cancelled":   "FFEBE6",
    "Blocked":     "FFEBE6",
}

def status_fill(status: str):
    color = STATUS_COLORS.get(status)
    return PatternFill("solid", start_color=color, end_color=color) if color else None

# (Header, row-key, Breite, h_align, wrap_text)
COLUMNS = [
    ("Team Mapping",       "team_mapped",        22, "center", False),
    ("Key",                "key",                18, "left",   False),
    ("Zusammenfassung",    "summary",            60, "left",   True),
    ("Functional Release", "functional_release", 22, "center", False),
    ("Release",            "release",            14, "center", False),
    ("Tracking\nRelease",  "tracking_release",   14, "center", False),
    ("Not-Testable",       "not_testable",       13, "center", False),
    ("Bugs\noffen",        "defects_open",       13, "center", False),
    ("Bugs\ngesamt",       "defects_total",      13, "center", False),
    ("Story\nInkonsistent", "inconsistent",    13, "center", False),
    ("CS\nInkonsistent",   "cs_inconsistent", 13, "center", False),
    ("Status",             "status",             22, "center", False),
]

INDENT_COLS = {"key", "summary"}

def export_excel(rows: list, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "CompStage Analyse"

    HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    CS_FILL     = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
    STORY_FILL  = PatternFill("solid", start_color="F5F9FF", end_color="F5F9FF")

    def make_border(color="BDD7EE"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    # --- Header ---
    for col_idx, (header, _, width, _, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        c.fill      = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = make_border("1F4E79")
        ws.column_dimensions[c.column_letter].width = width
    ws.row_dimensions[1].height = 30   # Platz für Zeilenumbruch im Header

    # --- Datenzeilen ---
    for row_idx, row in enumerate(rows, start=2):
        is_cs  = row["row_type"] == "compstage"
        fill   = CS_FILL if is_cs else STORY_FILL
        indent = 3 if not is_cs else 0

        for col_idx, (_, key, _, h_align, wrap) in enumerate(COLUMNS, start=1):
            value = row.get(key, "")
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.font   = Font(name="Arial", size=10, bold=is_cs)
            c.fill   = fill
            c.border = make_border()
            c.alignment = Alignment(
                horizontal = h_align,
                vertical   = "center",
                wrap_text  = wrap,
                indent     = indent if key in INDENT_COLS else 0,
            )

            # Key-Spalte: Hyperlink zu Jira
            if key == "key" and value:
                url = f"{JIRA_BASE_URL}/browse/{value}"
                c.hyperlink = url
                c.font = Font(
                    name="Arial", size=10, bold=is_cs,
                    color="0563C1", underline="single",
                )

            # Status: eigene Farbe
            if key == "status":
                sf = status_fill(value)
                if sf:
                    c.fill = sf

            # Tracking Release vs Release: rot wenn beide gesetzt und verschieden
            if key == "tracking_release" and value:
                release_val = row.get("release", "")
                if release_val and value != release_val:
                    c.fill = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC")
                    c.font = Font(name="Arial", size=10, bold=True, color="C00000")

            # Not-Testable: rote Schrift
            if key == "not_testable" and value == "X":
                c.font = Font(name="Arial", size=10, bold=True, color="C00000")

            # CS Inkonsistent (X auf CS-Zeile, x auf Story-Zeile)
            if key == "cs_inconsistent":
                if value == "X":   # CompStage selbst
                    c.fill = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC")
                    c.font = Font(name="Arial", size=10, bold=True, color="C00000")
                elif value == "x": # Story die Ursache ist → orange
                    c.fill = PatternFill("solid", start_color="FFE0B2", end_color="FFE0B2")
                    c.font = Font(name="Arial", size=10, bold=True, color="E65100")
            # Story Inkonsistent (offene Bugs bei closed Story)
            if key == "inconsistent" and value == "X":
                c.fill = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC")
                c.font = Font(name="Arial", size=10, bold=True, color="C00000")

            # Defects offen > 0: orange Hintergrund
            if key == "defects_open" and isinstance(value, int) and value > 0:
                c.fill = PatternFill("solid", start_color="FFE0B2", end_color="FFE0B2")
                c.font = Font(name="Arial", size=10, bold=True, color="E65100")

    ws.freeze_panes    = "A2"
    last_col_letter    = ws.cell(row=1, column=len(COLUMNS)).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(rows) + 1}"

    wb.save(path)
    print(f"\nExcel gespeichert : {path}")

# ---------------------------------------------------------------------------
# SQLite-Export mit Delta-Modus
# ---------------------------------------------------------------------------
def init_db(conn: sqlite3.Connection) -> None:
    """Erstellt Tabellen falls nicht vorhanden (Schema-Migration sicher)."""
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS compstages (
            key                 TEXT PRIMARY KEY,
            summary             TEXT,
            status              TEXT,
            team_mapped         TEXT,
            functional_release  TEXT,
            release             TEXT,
            not_testable        TEXT,
            cs_inconsistent     TEXT,
            fetched_at          DATETIME
        );

        CREATE TABLE IF NOT EXISTS stories (
            key            TEXT PRIMARY KEY,
            summary        TEXT,
            status         TEXT,
            release        TEXT,
            not_testable   TEXT,
            defects_open   INTEGER,
            defects_total  INTEGER,
            inconsistent   TEXT,
            fetched_at     DATETIME
        );

        -- n:m Verknüpfung CompStage ↔ Story
        CREATE TABLE IF NOT EXISTS compstage_stories (
            compstage_key  TEXT REFERENCES compstages(key),
            story_key      TEXT REFERENCES stories(key),
            PRIMARY KEY (compstage_key, story_key)
        );

        CREATE INDEX IF NOT EXISTS idx_cs_stories_cs    ON compstage_stories(compstage_key);
        CREATE INDEX IF NOT EXISTS idx_cs_stories_story ON compstage_stories(story_key);

        CREATE TABLE IF NOT EXISTS delta_log (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            run_at      DATETIME,
            issue_type  TEXT,        -- 'compstage' | 'story'
            issue_key   TEXT,
            change_type TEXT,        -- 'new' | 'updated' | 'unchanged'
            field       TEXT,        -- welches Feld hat sich geändert
            old_value   TEXT,
            new_value   TEXT
        );
    """)
    # Schema-Migrationen für bestehende DBs
    migrations = [
        "ALTER TABLE stories ADD COLUMN inconsistent TEXT DEFAULT ''",
        "ALTER TABLE stories DROP COLUMN parent_key",
        "ALTER TABLE compstages ADD COLUMN cs_inconsistent TEXT DEFAULT ''",
    ]
    for sql in migrations:
        try:
            conn.execute(sql)
        except Exception:
            pass
    conn.commit()


def _cs_tuple(row: dict) -> tuple:
    return (
        row["summary"], row["status"], row["team_mapped"],
        row["functional_release"], row["release"], row["not_testable"],
        row["cs_inconsistent"],
    )

def _st_tuple(row: dict) -> tuple:
    return (
        row["summary"], row["status"], row["release"],
        row["not_testable"], row["defects_open"], row["defects_total"],
        row["inconsistent"],
    )

CS_FIELDS_DB  = ["summary", "status", "team_mapped", "functional_release", "release", "not_testable", "cs_inconsistent"]
ST_FIELDS_DB  = ["summary", "status", "release", "not_testable", "defects_open", "defects_total", "inconsistent"]


def export_sqlite(rows: list, db_path: str) -> None:
    now     = datetime.now(timezone.utc).isoformat()
    conn    = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    init_db(conn)
    cur     = conn.cursor()

    # Bestehende Daten laden für Delta-Vergleich
    # Explizite Feldliste statt SELECT * – verhindert Reihenfolge-Probleme
    # nach ALTER TABLE Migrationen
    cs_select = "key, " + ", ".join(CS_FIELDS_DB)
    st_select = "key, " + ", ".join(ST_FIELDS_DB)
    existing_cs = {r["key"]: r for r in cur.execute(f"SELECT {cs_select} FROM compstages").fetchall()}
    existing_st = {r["key"]: r for r in cur.execute(f"SELECT {st_select} FROM stories").fetchall()}
    # Bestehende Verknüpfungen laden
    existing_links = set(
        (r[0], r[1]) for r in cur.execute("SELECT compstage_key, story_key FROM compstage_stories").fetchall()
    )
    new_links = set()

    delta_entries = []
    cs_new = cs_upd = cs_unch = 0
    st_new = st_upd = st_unch = 0

    def log(issue_type, key, change_type, field="", old="", new=""):
        delta_entries.append((now, issue_type, key, change_type, field, str(old), str(new)))

    for row in rows:
        if row["row_type"] == "compstage":
            key = row["key"]
            new_vals = _cs_tuple(row)

            if key not in existing_cs:
                cur.execute(
                    """INSERT INTO compstages
                       (key, summary, status, team_mapped, functional_release, release, not_testable, cs_inconsistent, fetched_at)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (key, *new_vals, now),
                )
                log("compstage", key, "new")
                cs_new += 1
            else:
                old_row  = existing_cs[key]
                old_vals = tuple(old_row[f] for f in CS_FIELDS_DB)
                if new_vals != old_vals:
                    for f, o, n in zip(CS_FIELDS_DB, old_vals, new_vals):
                        if o != n:
                            log("compstage", key, "updated", f, o, n)
                    cur.execute(
                        """UPDATE compstages
                           SET summary=?, status=?, team_mapped=?, functional_release=?,
                               release=?, not_testable=?, cs_inconsistent=?, fetched_at=?
                           WHERE key=?""",
                        (*new_vals, now, key),
                    )
                    cs_upd += 1
                else:
                    cs_unch += 1

        else:  # story – kann an mehreren CompStages hängen
            key        = row["key"]
            parent_key = row["parent_key"]
            new_vals   = _st_tuple(row)

            # Verknüpfung merken
            if parent_key:
                new_links.add((parent_key, key))

            # Story selbst nur einmal schreiben/aktualisieren
            if key not in existing_st:
                cur.execute(
                    """INSERT INTO stories
                       (key, summary, status, release, not_testable,
                        defects_open, defects_total, inconsistent, fetched_at)
                       VALUES (?,?,?,?,?,?,?,?,?)""",
                    (key, *new_vals, now),
                )
                log("story", key, "new")
                existing_st[key] = dict(zip(["key"] + ST_FIELDS_DB, [key] + list(new_vals)))
                st_new += 1
            else:
                old_row  = existing_st[key]
                old_vals = tuple(old_row[f] for f in ST_FIELDS_DB)
                if new_vals != old_vals:
                    for f, o, n in zip(ST_FIELDS_DB, old_vals, new_vals):
                        if str(o) != str(n):
                            log("story", key, "updated", f, o, n)
                    cur.execute(
                        """UPDATE stories
                           SET summary=?, status=?, release=?, not_testable=?,
                               defects_open=?, defects_total=?, inconsistent=?,
                               fetched_at=?
                           WHERE key=?""",
                        (*new_vals, now, key),
                    )
                    st_upd += 1
                else:
                    st_unch += 1

    # Neue Verknüpfungen in compstage_stories schreiben
    for link in new_links - existing_links:
        cur.execute(
            "INSERT OR IGNORE INTO compstage_stories (compstage_key, story_key) VALUES (?,?)",
            link,
        )

    # Delta-Log schreiben
    if delta_entries:
        cur.executemany(
            "INSERT INTO delta_log (run_at, issue_type, issue_key, change_type, field, old_value, new_value) "
            "VALUES (?,?,?,?,?,?,?)",
            delta_entries,
        )

    conn.commit()
    conn.close()

    print(f"SQLite gespeichert: {db_path}")
    print(f"  CompStages → {cs_new} neu · {cs_upd} geändert · {cs_unch} unverändert")
    print(f"  Stories    → {st_new} neu · {st_upd} geändert · {st_unch} unverändert")
    if delta_entries:
        changes = len([e for e in delta_entries if e[3] in ("new", "updated")])
        print(f"  Delta-Log  → {changes} Einträge geschrieben")

# ---------------------------------------------------------------------------
# Confluence-Export
# ---------------------------------------------------------------------------
def _cf_get(path: str, params: dict = None) -> dict:
    url  = f"{CONFLUENCE_BASE_URL}/rest/api{path}"
    resp = _cf_session.get(url, params=params or {})
    resp.raise_for_status()
    return resp.json()

def _cf_post(path: str, body: dict) -> dict:
    url  = f"{CONFLUENCE_BASE_URL}/rest/api{path}"
    resp = _cf_session.post(url, json=body)
    resp.raise_for_status()
    return resp.json()

def _cf_put(path: str, body: dict) -> dict:
    url  = f"{CONFLUENCE_BASE_URL}/rest/api{path}"
    resp = _cf_session.put(url, json=body)
    resp.raise_for_status()
    return resp.json()

def _cf_attach(page_id: str, file_path: str) -> None:
    url      = f"{CONFLUENCE_BASE_URL}/rest/api/content/{page_id}/child/attachment"
    filename = os.path.basename(file_path)
    # X-Atlassian-Token muss gesetzt sein, Content-Type darf NICHT manuell
    # gesetzt werden – requests setzt ihn automatisch korrekt als multipart/form-data
    s = requests.Session()
    s.headers.update({
        "Authorization":     f"Bearer {CONFLUENCE_PAT}",
        "X-Atlassian-Token": "no-check",
        "Accept":            "application/json",
    })
    s.verify = SSL_VERIFY
    with open(file_path, "rb") as f:
        resp = s.post(
            url,
            files={"file": (filename, f,
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"minorEdit": "true"},
        )
    if not resp.ok:
        print(f"  Anhang fehlgeschlagen ({resp.status_code}): {resp.text[:300]}")
    else:
        print(f"  Anhang hochgeladen: {filename}")

def _cf_find_page(title: str) -> dict | None:
    """Sucht Seite per Titel + Space, prüft Parent clientseitig.
    Confluence Server ignoriert den ancestors-GET-Parameter – daher manuelle Prüfung.
    """
    try:
        data = _cf_get("/content", {
            "spaceKey": CONFLUENCE_SPACE,
            "title":    title,
            "expand":   "version,ancestors",
            "limit":    25,
        })
        for page in data.get("results", []):
            ancestor_ids = {str(a["id"]) for a in page.get("ancestors", [])}
            if str(CONFLUENCE_PARENT_ID) in ancestor_ids:
                return page
        return None
    except Exception as e:
        print(f"  _cf_find_page Fehler: {e}")
        return None


def _build_confluence_body(run_date: str, rows: list,
                            db_summary: str, excel_filename: str,
                            page_id: str | None) -> str:
    cs_incons = sum(1 for r in rows if r["row_type"] == "compstage" and r.get("cs_inconsistent") == "X")
    st_incons = sum(1 for r in rows if r["row_type"] == "story"     and r.get("inconsistent")    == "X")
    cs_total  = sum(1 for r in rows if r["row_type"] == "compstage")
    st_total  = sum(1 for r in rows if r["row_type"] == "story")
    run_time  = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    # Excel-Anhang-Link (nur wenn Seite + Datei vorhanden)
    if page_id and excel_filename:
        att_url    = (f"{CONFLUENCE_BASE_URL}/download/attachments/{page_id}"
                      f"/{requests.utils.quote(excel_filename)}")
        excel_link = f'<a href="{att_url}">{html_mod.escape(excel_filename)}</a>'
    else:
        excel_link = html_mod.escape(excel_filename) if excel_filename else "–"

    # db_summary: nur die Zählzeilen, kein Pfad
    summary_lines = [l for l in db_summary.splitlines()
                     if "->" in l or "→" in l or "Delta-Log" in l]
    db_short = html_mod.escape(
        "\n".join(summary_lines).replace("→", "->")
    )

    ic = "color:rgb(200,0,0);font-weight:bold;"

    return (
        f'<p><strong>Lauf: {run_time}</strong></p>'
        f'<table><tbody>'
        f'<tr><td>CompStages gesamt</td><td>{cs_total}</td></tr>'
        f'<tr><td>Stories gesamt</td><td>{st_total}</td></tr>'
        f'<tr><td><span style="{ic}">CompStages inkonsistent</span></td>'
        f'<td><span style="{ic}">{cs_incons}</span></td></tr>'
        f'<tr><td><span style="{ic}">Stories inkonsistent (offene Bugs)</span></td>'
        f'<td><span style="{ic}">{st_incons}</span></td></tr>'
        f'</tbody></table>'
        f'<pre>{db_short}</pre>'
        f'<p>{excel_link}</p>'
    )


def export_confluence(rows: list, db_summary: str, excel_path: str) -> None:
    """Confluence-Export – Fehler werden abgefangen, Excel/DB sind nicht betroffen."""
    global _cf_session
    if not CONFLUENCE_PAT:
        print("⚠  CONFLUENCE_PAT nicht gesetzt – Confluence-Export übersprungen.")
        return

    _cf_session = requests.Session()
    _cf_session.headers.update({
        "Authorization": f"Bearer {CONFLUENCE_PAT}",
        "Accept":        "application/json",
        "Content-Type":  "application/json",
    })
    _cf_session.verify = SSL_VERIFY

    title          = datetime.now().strftime("%Y-%m-%d")
    excel_filename = os.path.basename(excel_path) if excel_path else ""

    try:
        existing = _cf_find_page(title)
    except Exception as e:
        print(f"⚠  Confluence: Seiten-Suche fehlgeschlagen: {e}")
        return

    if existing:
        page_id     = existing["id"]
        version_num = existing["version"]["number"] + 1
        print(f"  Confluence: bestehende Seite gefunden (ID={page_id}, v{version_num - 1})")
        body = _build_confluence_body(title, rows, db_summary, excel_filename, page_id)
        _cf_put(f"/content/{page_id}", {
            "version": {"number": version_num},
            "title":   title,
            "type":    "page",
            "body":    {"storage": {"value": body, "representation": "storage"}},
        })
        print(f"  Confluence: Seite aktualisiert")
    else:
        print(f"  Confluence: lege neue Seite an …")
        body      = _build_confluence_body(title, rows, db_summary, excel_filename, None)
        post_body = {
            "type":      "page",
            "title":     title,
            "space":     {"key": CONFLUENCE_SPACE},
            "ancestors": [{"id": CONFLUENCE_PARENT_ID}],
            "body":      {"storage": {"value": body, "representation": "storage"}},
        }
        post_url  = f"{CONFLUENCE_BASE_URL}/rest/api/content"
        # Debug: Body in Datei schreiben um problematische Zeichen zu finden
        debug_path = os.path.join(_output_dir, "confluence_body_debug.html")
        try:
            with open(debug_path, "w", encoding="utf-8") as _f:
                _f.write(body)
            print(f"  Body gespeichert: {debug_path}")
        except Exception:
            pass
        post_resp = _cf_session.post(post_url, json=post_body)
        if not post_resp.ok:
            print(f"  POST {post_resp.status_code}: {post_resp.text[:1000]}")
        post_resp.raise_for_status()
        page_id = post_resp.json()["id"]
        print(f"  Confluence: neue Seite erstellt (ID={page_id})")

    # Excel als Anhang hochladen
    if excel_path and os.path.exists(excel_path):
        _cf_attach(page_id, excel_path)
        # Seite nochmal updaten damit Anhang-Link korrekt ist
        try:
            ver         = _cf_get(f"/content/{page_id}", {"expand": "version"})
            version_num = ver["version"]["number"] + 1
            body        = _build_confluence_body(title, rows, db_summary, excel_filename, page_id)
            _cf_put(f"/content/{page_id}", {
                "version": {"number": version_num},
                "title":   title,
                "type":    "page",
                "body":    {"storage": {"value": body, "representation": "storage"}},
            })
        except Exception as e:
            print(f"  Anhang-Link Update fehlgeschlagen: {e}")

    page_url = f"{CONFLUENCE_BASE_URL}/pages/viewpage.action?pageId={page_id}"
    print(f"✓ Confluence    : {page_url}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("=" * 60)
    print("  Jira CompStage & Story Analyse")
    print(f"  Server  : {JIRA_BASE_URL}")
    print(f"  Projekt : {JIRA_PROJECT}")
    print(f"  SSL     : {'verify' if SSL_VERIFY else 'disabled (JIRA_SSL_VERIFY=false)'}")
    print(f"  DB      : {DB_PATH}")
    print("=" * 60 + "\n")

    print("Lade Tracking-Sheet …")
    tracking_map = load_tracking_releases()

    rows = build_rows(tracking_map)

    if not rows:
        print("Keine Daten – JQL oder Zugangsdaten prüfen.")
        sys.exit(1)

    cs_total   = sum(1 for r in rows if r["row_type"] == "compstage")
    st_total   = sum(1 for r in rows if r["row_type"] == "story")
    cs_incons  = sum(1 for r in rows if r["row_type"] == "compstage" and r.get("cs_inconsistent") == "X")
    st_incons  = sum(1 for r in rows if r["row_type"] == "story"     and r.get("inconsistent")    == "X")
    print(f"\nGesamt       : {cs_total} CompStages · {st_total} Stories")
    print(f"Inkonsistent : {cs_incons} CompStages · {st_incons} Stories")

    # Output-Verzeichnis anlegen falls nicht vorhanden
    os.makedirs(_output_dir, exist_ok=True)

    # 1. Excel (Haupt-Deliverable)
    try:
        export_excel(rows, OUTPUT_EXCEL)
        print(f"✓ Excel fertig : {OUTPUT_EXCEL}")
    except Exception as e:
        print(f"✗ Excel FEHLER : {e}")
        sys.exit(1)

    # 2. SQLite mit Delta-Modus – Summary für Confluence festhalten
    db_summary = ""
    print(f"\nSchreibe SQLite-DB (Delta-Modus) …")
    try:
        import io
        from contextlib import redirect_stdout
        buf = io.StringIO()
        with redirect_stdout(buf):
            export_sqlite(rows, DB_PATH)
        db_summary = buf.getvalue().strip()
        print(db_summary)
        print(f"✓ DB fertig    : {DB_PATH}")
    except Exception as e:
        db_summary = f"DB-Export fehlgeschlagen: {e}"
        print(f"⚠  {db_summary}")

    # 3. Confluence-Kurzzusammenfassung
    print(f"\nSchreibe Confluence-Seite …")
    try:
        export_confluence(rows, db_summary, OUTPUT_EXCEL)
    except Exception as e:
        print(f"⚠  Confluence-Export fehlgeschlagen: {e}")

    print("\n✓ Fertig")
